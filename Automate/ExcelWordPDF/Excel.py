import openpyxl, os
os.chdir(r"C:\Users\DavidLapadula\Downloads")

workbook = openpyxl.load_workbook("example.xlsx")
sheet = workbook.get_sheet_by_name("Sheet1")

a1 = sheet["A1"].value
b1 = sheet["B1"].value
c1 = sheet.cell(row=1, column=3).value

wb = openpyxl.Workbook()
sheet = wb.get_sheet_by_name("Sheet")

sheet["A1"] = 42
sheet["A2"] = "Hello"

sheet2 = wb.create_sheet(index=1, title="Using keyw argument for title")
sheet2.title("Some other sheet")
sheets = wb.get_sheet_names()
print(sheets)

wb.save("test.xlsx")

